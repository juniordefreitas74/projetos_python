# ======================================================
# GERADOR AUTOMÁTICO DE FOLHAS DE FREQUÊNCIA - ETEC
# ======================================================

# --------- IMPORTAR BIBLIOTECAS ---------
import time  # permite fazer o programa esperar alguns segundos
import pandas as pd  # biblioteca para ler arquivos Excel de forma fácil
from openpyxl import load_workbook  # biblioteca para editar arquivos Excel
from openpyxl.drawing.image import Image
import os  # biblioteca para mexer com pastas e arquivos
import shutil  # biblioteca para copiar arquivos
import win32com.client  # biblioteca para controlar o Excel pelo Python
from PyPDF2 import PdfMerger  # biblioteca para juntar vários PDFs em um só

BASE_DIR = os.path.dirname(os.path.abspath(__file__))


# função que espera o PDF ficar pronto de verdade
def esperar_pdf(caminho_pdf):
    tamanho_anterior = -1

    while True:
        if os.path.exists(caminho_pdf):
            tamanho_atual = os.path.getsize(caminho_pdf)

            # se o tamanho parou de crescer → terminou de gravar
            if tamanho_atual == tamanho_anterior and tamanho_atual > 0:
                break

            tamanho_anterior = tamanho_atual

        time.sleep(1)


# =============================
# 1 - CRIAR PASTA "folhas"
# =============================
# --------- CRIAR PASTA DE SAÍDA ---------
# se a pasta 'folhas' não existir criar ela
if not os.path.exists("folhas"):
    os.mkdir("folhas")

# LER PLANILHAS DOS PROFESSORES

dados = pd.read_excel("liskk.xlsx")
# abre a planilha dos professores

total_professores = len(dados)
# conta quantos professores existem
PAGINAS_RESERVAS = 5
# quantidade de paginas reservas em branco

TOTAL_PAGINAS = total_professores + PAGINAS_RESERVAS
# TOTAL DO LIVRO = professores + paginas extras
print("Professores encontrados: ", total_professores)
print("Total de páginas do livro: ", TOTAL_PAGINAS)

# ---------GERAR TODAS AS FOLHAS--------

for numero_pagina in range(1, TOTAL_PAGINAS + 1):
    # cria o nome do arquivo: 001.xlsx, 002.xlsx...
    nome_arquivo = f"folhas/{numero_pagina:03d}.xlsx"

    # copia o folhakk
    shutil.copy("folhakk.xlsx", nome_arquivo)

    # abre o arquivo copiado
    wb = load_workbook(nome_arquivo)
    ws = wb.active
    # img = Image(os.path.join(BASE_DIR, "cabecalho.png"))

    # img.anchor = "A1"  # posiciona na célula A1

    # ws.add_image(img)

    # sempre escreve o número da página
    ws["AD2"] = numero_pagina

    # se existir professor para essa página → preencher

    if numero_pagina <= total_professores:

        nome = dados.loc[numero_pagina - 1, "PROF"]
        matricula = dados.loc[numero_pagina - 1, "MATR"]
        sede = dados.loc[numero_pagina - 1, "SEDE"]

        ws["C4"] = nome  # escreve nome do professor
        ws["Q4"] = matricula  # escreve matrícula
        ws["V4"] = sede  # escreve sede
        print(f"Página {numero_pagina}: {nome}")
    else:
        # páginas extras ficam vazias
        print(f"Página {numero_pagina}: (em branco - reserva)")
    wb.save(nome_arquivo)  # salva arquivo
print("Arquivos Excel criados!")

# --------- CONVERTER TODOS PARA PDF ---------

print("Convertendo para PDF...")
excel = win32com.client.Dispatch("Excel.Application")
excel.Visible = False
pasta = os.path.abspath("folhas")
arquivos = sorted(os.listdir(pasta))
for arquivo in arquivos:
    if arquivo.endswith(".xlsx"):
        caminho = os.path
        caminho = os.path.join(pasta, arquivo)

        wb = excel.Workbooks.Open(caminho)

        pdf = caminho.replace(".xlsx", ".pdf")

        wb.ExportAsFixedFormat(0, pdf)
        wb.Close(False)

        # AGORA ESPERA O PDF TERMINAR DE SER CRIADO
        esperar_pdf(pdf)

        excel.Quit()


# --------- JUNTAR TODOS OS PDFs ---------
print("Juntando PDFs...")

merger = PdfMerger()

for arquivo in arquivos:
    if arquivo.endswith(".pdf"):
        merger.append(os.path.join(pasta, arquivo))

merger.write("FREQUENCIAS_FINAL.pdf")
merger.close()

print("\nPROCESSO FINALIZADO COM SUCESSO!")
