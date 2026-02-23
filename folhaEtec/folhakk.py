# ======================================================
# GERADOR AUTOMÁTICO DE FOLHAS DE FREQUÊNCIA - ETEC
# ======================================================

import time
import pandas as pd
from openpyxl import load_workbook
import os
import shutil
import win32com.client
from PyPDF2 import PdfMerger

BASE_DIR = os.path.dirname(os.path.abspath(__file__))


# ------------------------------------------------------
# FUNÇÃO PARA ESPERAR O PDF TERMINAR DE SER GERADO
# ------------------------------------------------------
def esperar_pdf(caminho_pdf):
    tamanho_anterior = -1

    while True:
        if os.path.exists(caminho_pdf):
            tamanho_atual = os.path.getsize(caminho_pdf)

            if tamanho_atual == tamanho_anterior and tamanho_atual > 0:
                break

            tamanho_anterior = tamanho_atual

        time.sleep(1)


# ------------------------------------------------------
# 1 - CRIAR / LIMPAR PASTA
# ------------------------------------------------------
pasta_saida = os.path.join(BASE_DIR, "folhakk")

if not os.path.exists(pasta_saida):
    os.mkdir(pasta_saida)

# apagar PDFs antigos
for arquivo in os.listdir(pasta_saida):
    if arquivo.lower().endswith(".pdf"):
        os.remove(os.path.join(pasta_saida, arquivo))


# ------------------------------------------------------
# 2 - LER PLANILHA DE PROFESSORES
# ------------------------------------------------------
dados = pd.read_excel(os.path.join(BASE_DIR, "liskk.xlsx"))

total_professores = len(dados)
PAGINAS_RESERVAS = 5
TOTAL_PAGINAS = total_professores + PAGINAS_RESERVAS

print("Professores encontrados:", total_professores)
print("Total de páginas do livro:", TOTAL_PAGINAS)


# ------------------------------------------------------
# 3 - GERAR ARQUIVOS EXCEL
# ------------------------------------------------------
for numero_pagina in range(1, TOTAL_PAGINAS + 1):

    nome_arquivo = os.path.join(pasta_saida, f"{numero_pagina:03d}.xlsx")

    shutil.copy(os.path.join(BASE_DIR, "folhakk.xlsx"), nome_arquivo)

    wb = load_workbook(nome_arquivo)
    ws = wb.active

    ws["AD2"] = numero_pagina

    if numero_pagina <= total_professores:

        nome = dados.loc[numero_pagina - 1, "PROF"]
        matricula = dados.loc[numero_pagina - 1, "MATR"]
        sede = dados.loc[numero_pagina - 1, "SEDE"]

        ws["C4"] = nome
        ws["Q4"] = matricula
        ws["V4"] = sede

        print(f"Página {numero_pagina}: {nome}")

    else:
        print(f"Página {numero_pagina}: (reserva)")

    wb.save(nome_arquivo)

print("Arquivos Excel criados!")


# ------------------------------------------------------
# 4 - CONVERTER TODOS PARA PDF
# ------------------------------------------------------
print("Convertendo para PDF...")

excel = win32com.client.DispatchEx("Excel.Application")
excel.Visible = False
excel.DisplayAlerts = False
excel.AutomationSecurity = 3  # desativa bloqueios

arquivos_xlsx = sorted(os.listdir(pasta_saida))

for arquivo in arquivos_xlsx:

    if not arquivo.lower().endswith(".xlsx"):
        continue

    caminho_xlsx = os.path.join(pasta_saida, arquivo)
    caminho_pdf = caminho_xlsx.replace(".xlsx", ".pdf")

    print("Gerando PDF:", arquivo.replace(".xlsx", ".pdf"))

    try:
        wb = excel.Workbooks.Open(caminho_xlsx)

        wb.ExportAsFixedFormat(
            Type=0,  # PDF
            Filename=caminho_pdf,
            Quality=0,
            IncludeDocProperties=True,
            IgnorePrintAreas=False,
            OpenAfterPublish=False,
        )

        wb.Close(False)

        esperar_pdf(caminho_pdf)

    except Exception as e:
        print("Erro ao converter:", arquivo)
        print(e)

excel.Quit()

print("Todos os PDFs foram criados!")


# ------------------------------------------------------
# 5 - JUNTAR TODOS OS PDFs
# ------------------------------------------------------
print("Juntando PDFs finais...")

merger = PdfMerger()

lista_pdfs = sorted(os.listdir(pasta_saida))

for arquivo in lista_pdfs:
    if arquivo.lower().endswith(".pdf"):
        caminho_pdf = os.path.join(pasta_saida, arquivo)
        print("Adicionando:", arquivo)
        merger.append(caminho_pdf)

saida_final = os.path.join(BASE_DIR, "FOLHA_KK_IMPRESSAO.pdf")

merger.write(saida_final)
merger.close()

print("\nPDF FINAL CRIADO COM SUCESSO!")
print(saida_final)
